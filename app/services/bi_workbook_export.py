"""
app/services/bi_workbook_export.py

Generate a multi-sheet Excel workbook (.xlsx) with charts and formatted
insight data suitable for Power BI, Excel, or any BI tool that supports
OOXML workbooks.

Sheets
------
1. Dashboard   — summary KPIs, confidence gauge, pipeline status
2. KPI Trends  — time-series metric data with line charts
3. Risk        — risk scores with bar chart
4. Forecasts   — forecast data with trend lines
5. Insights    — competitive analysis + strategic recommendations
"""

from __future__ import annotations

import io
import json
import logging
from collections import defaultdict
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from agent.nodes.node_result import payload_of
from app.services.bi_export_service import BIExportService, ExportResult

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
_HEADER_FONT = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_SUBHEADER_FONT = Font(name="Segoe UI", bold=True, size=10, color="2F5496")
_BODY_FONT = Font(name="Segoe UI", size=10)
_WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")


def _apply_header_row(ws: Any, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _apply_body_style(ws: Any, start_row: int, end_row: int, ncols: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER


def _auto_width(ws: Any, ncols: int, max_rows: int = 50) -> None:
    for col in range(1, ncols + 1):
        max_len = 0
        letter = get_column_letter(col)
        for row in range(1, min(max_rows + 1, ws.max_row + 1)):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = min(max(max_len + 3, 12), 50)


def _confidence_fill(score: float) -> PatternFill:
    if score >= 0.7:
        return _GREEN_FILL
    if score >= 0.4:
        return _YELLOW_FILL
    return _RED_FILL


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _build_dashboard_sheet(
    wb: Workbook,
    entity_name: str,
    business_type: str,
    pipeline_status: str,
    confidence_score: float,
    insight_payload: dict[str, Any],
    diagnostics: dict[str, Any],
) -> None:
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = "2F5496"

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = f"Insight Report — {entity_name}"
    title_cell.font = Font(name="Segoe UI", bold=True, size=16, color="2F5496")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font = Font(name="Segoe UI", size=9, color="808080")

    # Summary metrics table
    row = 4
    summary_data = [
        ("Entity", entity_name),
        ("Business Type", business_type),
        ("Pipeline Status", pipeline_status.upper()),
        ("Confidence Score", f"{confidence_score:.0%}"),
    ]

    # Signal integrity scores
    integrity_scores = diagnostics.get("signal_integrity_scores") or {}
    if integrity_scores:
        summary_data.append(("", ""))
        summary_data.append(("— Signal Integrity —", ""))
        for key, val in integrity_scores.items():
            label = key.replace("_", " ")
            summary_data.append((label, f"{val:.1%}" if isinstance(val, (int, float)) else str(val)))

    headers = ["Metric", "Value"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    _apply_header_row(ws, row, len(headers))

    for i, (label, value) in enumerate(summary_data):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).font = _SUBHEADER_FONT if label.startswith("—") else _BODY_FONT
        cell = ws.cell(row=r, column=2, value=value)
        cell.font = _BODY_FONT
        if label == "Confidence Score":
            cell.fill = _confidence_fill(confidence_score)
        if label == "Pipeline Status":
            if pipeline_status == "success":
                cell.fill = _GREEN_FILL
            elif pipeline_status == "partial":
                cell.fill = _YELLOW_FILL
            else:
                cell.fill = _RED_FILL

    _apply_body_style(ws, row + 1, row + len(summary_data), len(headers))

    # Competitive Analysis summary
    comp = insight_payload.get("competitive_analysis") or {}
    if comp:
        r = row + len(summary_data) + 3
        ws.merge_cells(f"A{r}:D{r}")
        ws.cell(row=r, column=1, value="Analysis Summary").font = Font(
            name="Segoe UI", bold=True, size=12, color="2F5496"
        )
        for label, key in [
            ("Summary", "summary"),
            ("Market Position", "market_position"),
            ("Relative Performance", "relative_performance"),
        ]:
            r += 1
            ws.cell(row=r, column=1, value=label).font = _SUBHEADER_FONT
            ws.merge_cells(f"B{r}:D{r}")
            cell = ws.cell(row=r, column=2, value=comp.get(key, ""))
            cell.font = _BODY_FONT
            cell.alignment = _WRAP_ALIGN

    _auto_width(ws, 4)
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30


def _build_kpi_sheet(wb: Workbook, kpi_result: ExportResult) -> None:
    ws = wb.create_sheet("KPI Trends")
    ws.sheet_properties.tabColor = "548235"

    if not kpi_result.rows:
        ws["A1"] = "No KPI data available."
        ws["A1"].font = _BODY_FONT
        return

    # Group by metric_name for charting
    metrics_by_name: dict[str, list[dict]] = defaultdict(list)
    for row in kpi_result.rows:
        name = row.get("metric_name")
        if name:
            metrics_by_name[name].append(row)

    # Write data table
    headers = ["period_end", "entity_name", "metric_name", "metric_value", "metric_unit"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _apply_header_row(ws, 1, len(headers))

    data_row = 2
    for row in kpi_result.rows:
        for c, h in enumerate(headers, 1):
            val = row.get(h)
            cell = ws.cell(row=data_row, column=c, value=val)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
        data_row += 1

    _auto_width(ws, len(headers))

    # Build one line chart per metric
    chart_col = len(headers) + 2
    chart_row = 1
    for metric_name, metric_rows in metrics_by_name.items():
        # Sort by period_end
        sorted_rows = sorted(metric_rows, key=lambda r: str(r.get("period_end") or ""))

        if len(sorted_rows) < 2:
            continue

        # Write mini-table for chart data (dates + values)
        start_col = chart_col
        ws.cell(row=chart_row, column=start_col, value="Date")
        ws.cell(row=chart_row, column=start_col + 1, value=metric_name)
        for i, r in enumerate(sorted_rows):
            period = str(r.get("period_end") or "")[:10]
            value = r.get("metric_value")
            ws.cell(row=chart_row + 1 + i, column=start_col, value=period)
            ws.cell(row=chart_row + 1 + i, column=start_col + 1, value=value)

        n = len(sorted_rows)
        chart = LineChart()
        chart.title = metric_name.replace("_", " ").title()
        chart.style = 10
        chart.y_axis.title = metric_name
        chart.x_axis.title = "Period"
        chart.width = 20
        chart.height = 12

        data_ref = Reference(ws, min_col=start_col + 1, min_row=chart_row, max_row=chart_row + n)
        cats_ref = Reference(ws, min_col=start_col, min_row=chart_row + 1, max_row=chart_row + n)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.line.width = 25000

        anchor = f"{get_column_letter(start_col + 3)}{chart_row}"
        ws.add_chart(chart, anchor)
        chart_row += max(n + 3, 18)


def _build_risk_sheet(wb: Workbook, risk_result: ExportResult) -> None:
    ws = wb.create_sheet("Risk Assessment")
    ws.sheet_properties.tabColor = "BF0000"

    if not risk_result.rows:
        ws["A1"] = "No risk data available."
        ws["A1"].font = _BODY_FONT
        return

    headers = ["entity_name", "period_end", "risk_score"]
    extra_keys = [k for k in (risk_result.fields or []) if k.startswith("risk_meta__")]
    all_headers = headers + extra_keys[:10]

    for c, h in enumerate(all_headers, 1):
        ws.cell(row=1, column=c, value=h.replace("risk_meta__", ""))
    _apply_header_row(ws, 1, len(all_headers))

    for i, row in enumerate(risk_result.rows):
        for c, h in enumerate(all_headers, 1):
            val = row.get(h)
            cell = ws.cell(row=2 + i, column=c, value=val)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER

    _apply_body_style(ws, 2, 1 + len(risk_result.rows), len(all_headers))
    _auto_width(ws, len(all_headers))

    # Risk score bar chart
    if len(risk_result.rows) >= 1:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Risk Scores Over Time"
        chart.style = 10
        chart.y_axis.title = "Risk Score"
        chart.width = 20
        chart.height = 12

        n = len(risk_result.rows)
        data_ref = Reference(ws, min_col=3, min_row=1, max_row=1 + n)
        cats_ref = Reference(ws, min_col=2, min_row=2, max_row=1 + n)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        ws.add_chart(chart, f"{get_column_letter(len(all_headers) + 2)}1")


def _build_forecast_sheet(wb: Workbook, forecast_result: ExportResult) -> None:
    ws = wb.create_sheet("Forecasts")
    ws.sheet_properties.tabColor = "ED7D31"

    if not forecast_result.rows:
        ws["A1"] = "No forecast data available."
        ws["A1"].font = _BODY_FONT
        return

    # Identify forecast month columns
    all_fields = forecast_result.fields or []
    base_headers = ["entity_name", "metric_name", "period_end", "forecast__slope", "forecast__trend"]
    month_cols = sorted([f for f in all_fields if f.startswith("forecast__month_")])
    headers = base_headers + month_cols

    for c, h in enumerate(headers, 1):
        label = h.replace("forecast__", "").replace("_", " ").title()
        ws.cell(row=1, column=c, value=label)
    _apply_header_row(ws, 1, len(headers))

    for i, row in enumerate(forecast_result.rows):
        for c, h in enumerate(headers, 1):
            val = row.get(h)
            cell = ws.cell(row=2 + i, column=c, value=val)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER

    _auto_width(ws, len(headers))

    # Line chart for each metric's forecast months
    if month_cols and len(forecast_result.rows) >= 1:
        chart_row = len(forecast_result.rows) + 4
        for i, row in enumerate(forecast_result.rows[:5]):
            metric = row.get("metric_name") or f"Metric {i}"
            month_values = [row.get(m) for m in month_cols]
            if not any(v is not None for v in month_values):
                continue

            start_col = 1
            ws.cell(row=chart_row, column=start_col, value="Month")
            ws.cell(row=chart_row, column=start_col + 1, value=str(metric))
            for j, mc in enumerate(month_cols):
                ws.cell(row=chart_row + 1 + j, column=start_col, value=mc.replace("forecast__", ""))
                ws.cell(row=chart_row + 1 + j, column=start_col + 1, value=row.get(mc))

            n = len(month_cols)
            chart = LineChart()
            chart.title = f"Forecast — {metric}"
            chart.style = 10
            chart.width = 20
            chart.height = 12

            data_ref = Reference(ws, min_col=start_col + 1, min_row=chart_row, max_row=chart_row + n)
            cats_ref = Reference(ws, min_col=start_col, min_row=chart_row + 1, max_row=chart_row + n)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)

            ws.add_chart(chart, f"D{chart_row}")
            chart_row += max(n + 3, 18)


def _build_insights_sheet(wb: Workbook, insight_payload: dict[str, Any]) -> None:
    ws = wb.create_sheet("Strategic Insights")
    ws.sheet_properties.tabColor = "7030A0"

    comp = insight_payload.get("competitive_analysis") or {}
    strat = insight_payload.get("strategic_recommendations") or {}

    row = 1
    ws.merge_cells("A1:C1")
    ws["A1"].value = "Competitive Analysis"
    ws["A1"].font = Font(name="Segoe UI", bold=True, size=14, color="7030A0")

    row = 3
    analysis_fields = [
        ("Summary", comp.get("summary", "")),
        ("Market Position", comp.get("market_position", "")),
        ("Relative Performance", comp.get("relative_performance", "")),
        ("Confidence", f"{comp.get('confidence', 0):.0%}"),
    ]

    advantages = comp.get("key_advantages") or []
    vulnerabilities = comp.get("key_vulnerabilities") or []

    for label, value in analysis_fields:
        ws.cell(row=row, column=1, value=label).font = _SUBHEADER_FONT
        ws.merge_cells(f"B{row}:C{row}")
        cell = ws.cell(row=row, column=2, value=str(value))
        cell.font = _BODY_FONT
        cell.alignment = _WRAP_ALIGN
        row += 1

    if advantages:
        row += 1
        ws.cell(row=row, column=1, value="Key Advantages").font = _SUBHEADER_FONT
        for adv in advantages:
            row += 1
            ws.cell(row=row, column=2, value=f"• {adv}").font = _BODY_FONT
            ws.cell(row=row, column=2).alignment = _WRAP_ALIGN

    if vulnerabilities:
        row += 1
        ws.cell(row=row, column=1, value="Key Vulnerabilities").font = _SUBHEADER_FONT
        for vuln in vulnerabilities:
            row += 1
            ws.cell(row=row, column=2, value=f"• {vuln}").font = _BODY_FONT
            ws.cell(row=row, column=2).alignment = _WRAP_ALIGN

    # Strategic Recommendations
    row += 2
    ws.merge_cells(f"A{row}:C{row}")
    ws.cell(row=row, column=1, value="Strategic Recommendations").font = Font(
        name="Segoe UI", bold=True, size=14, color="7030A0"
    )

    sections = [
        ("Immediate Actions", strat.get("immediate_actions") or []),
        ("Mid-Term Moves", strat.get("mid_term_moves") or []),
        ("Defensive Strategies", strat.get("defensive_strategies") or []),
        ("Offensive Strategies", strat.get("offensive_strategies") or []),
    ]

    for section_title, items in sections:
        row += 2
        ws.cell(row=row, column=1, value=section_title).font = _SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = PatternFill(start_color="E8E0F0", end_color="E8E0F0", fill_type="solid")
        for item in items:
            row += 1
            ws.merge_cells(f"B{row}:C{row}")
            ws.cell(row=row, column=2, value=f"→ {item}").font = _BODY_FONT
            ws.cell(row=row, column=2).alignment = _WRAP_ALIGN

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 30


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def generate_bi_workbook(
    db: Session,
    *,
    entity_name: str,
    business_type: str,
    graph_state: dict[str, Any],
) -> bytes:
    """Build a multi-sheet .xlsx workbook from the pipeline output.

    Returns the workbook as raw bytes suitable for streaming.
    """
    # Extract data from graph state
    final_response = graph_state.get("final_response") or "{}"
    try:
        insight_payload = json.loads(final_response) if isinstance(final_response, str) else {}
    except (json.JSONDecodeError, TypeError):
        insight_payload = {}

    pipeline_status = str(graph_state.get("pipeline_status") or "unknown")
    diagnostics = graph_state.get("envelope_diagnostics") or {}
    confidence_score = float(diagnostics.get("confidence_score", 0))
    comp = insight_payload.get("competitive_analysis") or {}
    if confidence_score == 0 and comp.get("confidence"):
        confidence_score = float(comp["confidence"])

    # Fetch tabular data from the export service
    service = BIExportService()
    kpi_result = service.export(db, dataset="kpis", entity_name=entity_name)
    risk_result = service.export(db, dataset="risk", entity_name=entity_name)
    forecast_result = service.export(db, dataset="forecasts", entity_name=entity_name)

    # Build workbook
    wb = Workbook()

    _build_dashboard_sheet(
        wb,
        entity_name=entity_name,
        business_type=business_type,
        pipeline_status=pipeline_status,
        confidence_score=confidence_score,
        insight_payload=insight_payload,
        diagnostics=diagnostics,
    )
    _build_kpi_sheet(wb, kpi_result)
    _build_risk_sheet(wb, risk_result)
    _build_forecast_sheet(wb, forecast_result)
    _build_insights_sheet(wb, insight_payload)

    # Serialize to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
